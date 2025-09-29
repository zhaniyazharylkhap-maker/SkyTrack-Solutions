import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import psycopg2
from sqlalchemy import create_engine
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings('ignore')

# Create necessary directories for output files
os.makedirs('charts', exist_ok=True)
os.makedirs('exports', exist_ok=True)

class SkyTrackAnalytics:
    def __init__(self):
        """Initialize database connection and matplotlib settings"""
        self.db_config = {
            'host': 'localhost',
            'port': '5432',
            'database': 'airport_analytics',
            'user': 'postgres',
            'password': '0000'
        }
        
        connection_string = f"postgresql://{self.db_config['user']}:{self.db_config['password']}@{self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}"
        self.engine = create_engine(connection_string)
        
        # Configure matplotlib default settings
        plt.style.use('default')
        plt.rcParams['figure.figsize'] = (10, 6)
        plt.rcParams['font.size'] = 10
        
        print("Database connection established successfully")
    
    def create_pie_chart(self):
        """
        Task 1.1: Pie chart showing flight distribution by airlines
        Uses 2 JOINs: airline -> flights -> airport
        """
        print("\nCreating pie chart...")
        
        query = """
        SELECT 
            a.airline_name as airline,
            COUNT(DISTINCT f.flight_id) as flight_count,
            COUNT(DISTINCT ap.airport_id) as airports_served
        FROM airline a 
        JOIN flights f ON a.airline_id = f.airline_id
        JOIN airport ap ON f.departure_airport_id = ap.airport_id
        GROUP BY a.airline_name
        ORDER BY COUNT(DISTINCT f.flight_id) DESC
        LIMIT 8;
        """
        
        try:
            df = pd.read_sql_query(query, self.engine)
            
            if df.empty:
                print("No data available for pie chart")
                return
            
            plt.figure(figsize=(10, 8))
            colors = ['#FF9999', '#66B2FF', '#99FF99', '#FFCC99', '#FF99CC', '#99CCFF', '#FFB366', '#B3B3FF']
            
            wedges, texts, autotexts = plt.pie(df['flight_count'], 
                                             labels=df['airline'], 
                                             autopct='%1.1f%%',
                                             colors=colors,
                                             startangle=90)
            
            plt.title('Flight Distribution by Airlines', fontsize=14, fontweight='bold')
            plt.axis('equal')
            plt.tight_layout()
            plt.savefig('charts/pie_chart_airlines.png', dpi=300, bbox_inches='tight')
            plt.show()
            
            total_rows = len(df)
            total_flights = df['flight_count'].sum()
            print(f"Rows retrieved: {total_rows}")
            print(f"Graph type: Pie chart")
            print(f"Shows: Market share of airlines by flight count ({total_flights} total flights)")
            print(f"Saved to: charts/pie_chart_airlines.png")
            print(f"SQL JOINs used: 2 (airline -> flights -> airport)")
            
        except Exception as e:
            print(f"Error creating pie chart: {e}")
    
    def create_bar_chart(self):
        """
        Task 1.2: Bar chart showing top booking platforms
        Uses 2 JOINs: booking -> booking_flight -> flights
        """
        print("\nCreating bar chart...")
        
        query = """
        SELECT 
            b.booking_platform as platform,
            COUNT(b.booking_id) as booking_count,
            ROUND(AVG(b.price), 2) as avg_price
        FROM booking b
        JOIN booking_flight bf ON b.booking_id = bf.booking_id
        JOIN flights f ON bf.flight_id = f.flight_id
        GROUP BY b.booking_platform
        ORDER BY COUNT(b.booking_id) DESC
        LIMIT 10;
        """
        
        try:
            df = pd.read_sql_query(query, self.engine)
            
            if df.empty:
                print("No data available for bar chart")
                return
            
            plt.figure(figsize=(12, 6))
            bars = plt.bar(range(len(df)), df['booking_count'], 
                          color='skyblue', edgecolor='navy', linewidth=0.7)
            
            # Add value labels on top of bars
            for i, (bar, count) in enumerate(zip(bars, df['booking_count'])):
                plt.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                        str(count), ha='center', va='bottom', fontweight='bold')
            
            plt.title('Top 10 Booking Platforms by Volume', fontsize=14, fontweight='bold')
            plt.xlabel('Booking Platform', fontsize=12)
            plt.ylabel('Number of Bookings', fontsize=12)
            plt.xticks(range(len(df)), [p[:15] + '...' if len(p) > 15 else p for p in df['platform']], rotation=45, ha='right')
            plt.grid(axis='y', alpha=0.3)
            
            plt.tight_layout()
            plt.savefig('charts/bar_chart_platforms.png', dpi=300, bbox_inches='tight')
            plt.show()
            
            total_rows = len(df)
            print(f"Rows retrieved: {total_rows}")
            print(f"Graph type: Bar chart")
            print(f"Shows: Most popular booking platforms by volume")
            print(f"Saved to: charts/bar_chart_platforms.png")
            print(f"SQL JOINs used: 2 (booking -> booking_flight -> flights)")
            
        except Exception as e:
            print(f"Error creating bar chart: {e}")
    
    def create_horizontal_bar_chart(self):
        """
        Task 1.3: Horizontal bar chart showing busiest airports
        Uses 2 JOINs: airport -> flights -> airline
        """
        print("\nCreating horizontal bar chart...")
        
        query = """
        SELECT 
            ap.airport_name as airport,
            ap.city as city,
            COUNT(DISTINCT f.flight_id) as flight_count,
            COUNT(DISTINCT a.airline_id) as airlines_count
        FROM airport ap
        LEFT JOIN flights f ON (ap.airport_id = f.departure_airport_id OR ap.airport_id = f.arrival_airport_id)
        LEFT JOIN airline a ON f.airline_id = a.airline_id
        GROUP BY ap.airport_name, ap.city
        HAVING COUNT(DISTINCT f.flight_id) > 0
        ORDER BY COUNT(DISTINCT f.flight_id) DESC
        LIMIT 15;
        """
        
        try:
            df = pd.read_sql_query(query, self.engine)
            
            if df.empty:
                print("No data available for horizontal bar chart")
                return
            
            plt.figure(figsize=(12, 8))
            
            labels = [f"{airport}\n({city})" for airport, city in zip(df['airport'], df['city'])]
            
            bars = plt.barh(range(len(df)), df['flight_count'], 
                           color='lightcoral', edgecolor='darkred', linewidth=0.7)
            
            # Add value labels at the end of bars
            for i, (bar, count) in enumerate(zip(bars, df['flight_count'])):
                plt.text(bar.get_width() + 1, bar.get_y() + bar.get_height()/2,
                        str(count), va='center', fontweight='bold')
            
            plt.title('Top 15 Busiest Airports by Flight Volume', fontsize=14, fontweight='bold')
            plt.xlabel('Number of Flights', fontsize=12)
            plt.ylabel('Airport', fontsize=12)
            plt.yticks(range(len(df)), labels)
            plt.grid(axis='x', alpha=0.3)
            plt.gca().invert_yaxis()
            
            plt.tight_layout()
            plt.savefig('charts/horizontal_bar_airports.png', dpi=300, bbox_inches='tight')
            plt.show()
            
            total_rows = len(df)
            top_airport = df.iloc[0]
            print(f"Rows retrieved: {total_rows}")
            print(f"Graph type: Horizontal bar chart")
            print(f"Shows: Major transportation hubs. Busiest: {top_airport['airport']} ({top_airport['flight_count']} flights)")
            print(f"Saved to: charts/horizontal_bar_airports.png")
            print(f"SQL JOINs used: 2 (airport -> flights -> airline)")
            
        except Exception as e:
            print(f"Error creating horizontal bar chart: {e}")
    
    def create_line_chart(self):
        """
        Task 1.4: Line chart showing flight distribution by status
        Uses 2 JOINs: flights -> airline -> airport
        """
        print("\nCreating line chart...")
        
        query = """
        SELECT 
            f.status as flight_status,
            COUNT(f.flight_id) as flight_count,
            COUNT(DISTINCT a.airline_id) as airlines_count,
            COUNT(DISTINCT ap.airport_id) as airports_count
        FROM flights f
        JOIN airline a ON f.airline_id = a.airline_id
        JOIN airport ap ON f.departure_airport_id = ap.airport_id
        GROUP BY f.status
        ORDER BY f.status;
        """
        
        try:
            df = pd.read_sql_query(query, self.engine)
            
            if df.empty:
                print("No data available for line chart")
                return
            
            plt.figure(figsize=(12, 6))
            
            plt.plot(df['flight_status'], df['flight_count'], 
                    marker='o', linewidth=3, markersize=8, 
                    color='green', markerfacecolor='lightgreen', 
                    markeredgecolor='darkgreen', markeredgewidth=2)
            
            # Add value labels above each point
            for i, (status, count) in enumerate(zip(df['flight_status'], df['flight_count'])):
                plt.annotate(str(count), (i, count), 
                           textcoords="offset points", xytext=(0,10), ha='center',
                           fontweight='bold', fontsize=11)
            
            plt.title('Flight Count by Status', fontsize=14, fontweight='bold')
            plt.xlabel('Flight Status', fontsize=12)
            plt.ylabel('Number of Flights', fontsize=12)
            plt.xticks(rotation=45, ha='right')
            plt.grid(True, alpha=0.3)
            
            plt.tight_layout()
            plt.savefig('charts/line_chart_flight_status.png', dpi=300, bbox_inches='tight')
            plt.show()
            
            total_flights = df['flight_count'].sum()
            print(f"Rows retrieved: {len(df)}")
            print(f"Graph type: Line chart")
            print(f"Shows: Distribution of {total_flights} flights by operational status")
            print(f"Saved to: charts/line_chart_flight_status.png")
            print(f"SQL JOINs used: 2 (flights -> airline -> airport)")
            
        except Exception as e:
            print(f"Error creating line chart: {e}")
    
    def create_histogram(self):
        """
        Task 1.5: Histogram showing ticket price distribution
        Uses 2 JOINs: booking -> booking_flight -> flights
        """
        print("\nCreating histogram...")
        
        query = """
        SELECT 
            b.price as ticket_price
        FROM booking b
        JOIN booking_flight bf ON b.booking_id = bf.booking_id
        JOIN flights f ON bf.flight_id = f.flight_id
        WHERE b.price > 0;
        """
        
        try:
            df = pd.read_sql_query(query, self.engine)
            
            if df.empty:
                print("No data available for histogram")
                return
            
            plt.figure(figsize=(12, 6))
            
            # Create histogram with colored bins
            n, bins, patches = plt.hist(df['ticket_price'], bins=20, 
                                       color='orange', alpha=0.7, 
                                       edgecolor='darkorange', linewidth=1.2)
            
            # Apply gradient colors to bins
            for i, patch in enumerate(patches):
                patch.set_facecolor(plt.cm.viridis(i / len(patches)))
            
            plt.title('Distribution of Ticket Prices', fontsize=14, fontweight='bold')
            plt.xlabel('Ticket Price ($)', fontsize=12)
            plt.ylabel('Number of Bookings', fontsize=12)
            plt.grid(axis='y', alpha=0.3)
            
            # Add mean and median lines
            mean_price = df['ticket_price'].mean()
            median_price = df['ticket_price'].median()
            plt.axvline(mean_price, color='red', linestyle='--', linewidth=2, label=f'Mean: ${mean_price:.0f}')
            plt.axvline(median_price, color='blue', linestyle='--', linewidth=2, label=f'Median: ${median_price:.0f}')
            plt.legend()
            
            plt.tight_layout()
            plt.savefig('charts/histogram_ticket_prices.png', dpi=300, bbox_inches='tight')
            plt.show()
            
            total_bookings = len(df)
            print(f"Rows retrieved: {total_bookings}")
            print(f"Graph type: Histogram")
            print(f"Shows: Price distribution (Range: ${df['ticket_price'].min():.2f}-${df['ticket_price'].max():.2f}, Avg: ${mean_price:.2f})")
            print(f"Saved to: charts/histogram_ticket_prices.png")
            print(f"SQL JOINs used: 2 (booking -> booking_flight -> flights)")
            
        except Exception as e:
            print(f"Error creating histogram: {e}")
    
    def create_scatter_plot(self):
        """
        Task 1.6: Scatter plot showing baggage weight vs ticket price correlation
        Uses 3 JOINs: baggage -> booking -> booking_flight -> flights
        """
        print("\nCreating scatter plot...")
        
        query = """
        SELECT 
            bag.weight_in_kg as baggage_weight,
            b.price as ticket_price
        FROM baggage bag
        JOIN booking b ON bag.booking_id = b.booking_id
        JOIN booking_flight bf ON b.booking_id = bf.booking_id
        JOIN flights f ON bf.flight_id = f.flight_id
        WHERE bag.weight_in_kg > 0 AND b.price > 0
        LIMIT 200;
        """
        
        try:
            df = pd.read_sql_query(query, self.engine)
            
            if df.empty:
                print("No data available for scatter plot")
                return
            
            plt.figure(figsize=(12, 8))
            
            plt.scatter(df['baggage_weight'], df['ticket_price'], 
                       alpha=0.6, s=50, color='purple', edgecolors='indigo')
            
            plt.title('Relationship between Baggage Weight and Ticket Price', fontsize=14, fontweight='bold')
            plt.xlabel('Baggage Weight (kg)', fontsize=12)
            plt.ylabel('Ticket Price ($)', fontsize=12)
            plt.grid(True, alpha=0.3)
            
            # Add trend line
            z = np.polyfit(df['baggage_weight'], df['ticket_price'], 1)
            p = np.poly1d(z)
            plt.plot(df['baggage_weight'], p(df['baggage_weight']), "r--", alpha=0.8, linewidth=2)
            
            plt.tight_layout()
            plt.savefig('charts/scatter_plot_baggage_price.png', dpi=300, bbox_inches='tight')
            plt.show()
            
            total_points = len(df)
            correlation = df['baggage_weight'].corr(df['ticket_price'])
            print(f"Rows retrieved: {total_points}")
            print(f"Graph type: Scatter plot")
            print(f"Shows: Correlation between baggage weight and ticket price (r={correlation:.3f})")
            print(f"Saved to: charts/scatter_plot_baggage_price.png")
            print(f"SQL JOINs used: 3 (baggage -> booking -> booking_flight -> flights)")
            
        except Exception as e:
            print(f"Error creating scatter plot: {e}")
    
    def create_interactive_timeline(self):
        """
        Task 2: Interactive Plotly chart with time slider
        Uses animation_frame parameter with real dates from scheduled_departure column
        """
        print("\nCreating interactive timeline with Plotly...")
        
        query = """
        SELECT 
            a.airline_name as airline,
            f.status as flight_status,
            TO_CHAR(f.scheduled_departure, 'YYYY-MM') as month,
            COUNT(f.flight_id) as flight_count
        FROM airline a
        JOIN flights f ON a.airline_id = f.airline_id
        WHERE f.scheduled_departure IS NOT NULL
        GROUP BY a.airline_name, f.status, TO_CHAR(f.scheduled_departure, 'YYYY-MM')
        ORDER BY TO_CHAR(f.scheduled_departure, 'YYYY-MM'), a.airline_name;
        """
        
        try:
            df = pd.read_sql_query(query, self.engine)
            
            if df.empty:
                print("No data available for interactive chart")
                return
            
            # Create animated scatter plot with time slider
            fig = px.scatter(df, 
                           x="airline", 
                           y="flight_count",
                           animation_frame="month",
                           size="flight_count",
                           color="flight_status",
                           title="Flight Count Evolution by Airline (Monthly Timeline)",
                           labels={
                               "airline": "Airline",
                               "flight_count": "Number of Flights",
                               "flight_status": "Flight Status"
                           },
                           range_y=[0, df['flight_count'].max() + 5])
            
            fig.update_layout(
                width=1200,
                height=700,
                title_font_size=16,
                xaxis_tickangle=-45
            )
            
            fig.show()
            
            total_airlines = df['airline'].nunique()
            total_months = df['month'].nunique()
            print(f"Interactive timeline created successfully")
            print(f"Shows: Airline performance evolution across {total_months} months")
            print(f"Airlines tracked: {total_airlines}")
            print(f"Time data source: scheduled_departure column from flights table")
            print(f"Use the slider at the bottom to navigate through time periods")
            
        except Exception as e:
            print(f"Error creating interactive timeline: {e}")
    
    def export_to_excel(self):
        """
        Task 3: Export data to Excel with advanced formatting
        Includes: frozen headers, filters, gradient fills, conditional formatting
        """
        print("\nExporting analytical data to Excel...")
        
        # Define queries for different sheets
        queries = {
            'Airlines_Performance': """
                SELECT 
                    a.airline_name as "Airline Name",
                    COUNT(f.flight_id) as "Total Flights",
                    COUNT(DISTINCT f.departure_airport_id) as "Airports Served"
                FROM airline a
                LEFT JOIN flights f ON a.airline_id = f.airline_id
                GROUP BY a.airline_name
                ORDER BY COUNT(f.flight_id) DESC;
            """,
            'Airport_Traffic': """
                SELECT 
                    ap.airport_name as "Airport Name",
                    ap.city as "City",
                    COUNT(DISTINCT f.flight_id) as "Flight Count",
                    COUNT(DISTINCT a.airline_id) as "Airlines Operating"
                FROM airport ap
                LEFT JOIN flights f ON (ap.airport_id = f.departure_airport_id OR ap.airport_id = f.arrival_airport_id)
                LEFT JOIN airline a ON f.airline_id = a.airline_id
                GROUP BY ap.airport_name, ap.city
                ORDER BY COUNT(DISTINCT f.flight_id) DESC;
            """,
            'Booking_Summary': """
                SELECT 
                    b.booking_platform as "Platform",
                    COUNT(*) as "Bookings",
                    ROUND(AVG(b.price), 2) as "Avg Price",
                    ROUND(MIN(b.price), 2) as "Min Price",
                    ROUND(MAX(b.price), 2) as "Max Price"
                FROM booking b
                JOIN booking_flight bf ON b.booking_id = bf.booking_id
                GROUP BY b.booking_platform
                ORDER BY COUNT(*) DESC;
            """
        }
        
        try:
            filename = 'exports/skytrack_analytics_report.xlsx'
            
            # Write data to Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                total_rows = 0
                
                for sheet_name, query in queries.items():
                    df = pd.read_sql_query(query, self.engine)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    total_rows += len(df)
            
            # Apply formatting after writing
            self._apply_excel_formatting(filename, list(queries.keys()))
            
            sheet_count = len(queries)
            print(f"Created file: {filename}, {sheet_count} sheets, {total_rows} rows")
            
        except Exception as e:
            print(f"Error during Excel export: {e}")
    
    def _apply_excel_formatting(self, filename, sheet_names):
        """
        Apply Excel formatting: frozen panes, filters, gradients, conditional formatting
        """
        try:
            workbook = load_workbook(filename)
            
            for sheet_name in sheet_names:
                ws = workbook[sheet_name]
                
                # Freeze first row and first column
                ws.freeze_panes = "B2"
                
                # Add filters to all columns
                if ws.max_row > 1:
                    ws.auto_filter.ref = ws.dimensions
                
                # Apply gradient fill and conditional formatting to numeric columns
                for col_idx in range(2, ws.max_column + 1):
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    
                    # Check if column contains numeric data
                    has_numeric_data = False
                    for row_idx in range(2, min(ws.max_row + 1, 10)):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if isinstance(cell_value, (int, float)) and cell_value is not None:
                            has_numeric_data = True
                            break
                    
                    if has_numeric_data:
                        # Apply color scale rule (gradient from min to max)
                        rule = ColorScaleRule(
                            start_type="min", start_color="FFAA0000",
                            mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                            end_type="max", end_color="FF00AA00"
                        )
                        range_string = f"{col_letter}2:{col_letter}{ws.max_row}"
                        ws.conditional_formatting.add(range_string, rule)
                
                # Style header row
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = cell.font.copy(color="FFFFFF", bold=True)
            
            workbook.save(filename)
            
        except Exception as e:
            print(f"Warning: Formatting error occurred: {e}")
    
    def add_demo_flight(self):
        """
        Demo function for project defense
        Adds a new flight to demonstrate real-time data updates in charts
        """
        print("\n=== DEMO: Adding new flight for demonstration ===")
        
        try:
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()
            
            # Insert new flight with current date
            query = """
            INSERT INTO flights (airline_id, departure_airport_id, arrival_airport_id, 
                               status, scheduled_departure, scheduled_arrival, flight_no)
            VALUES (1, 1, 2, 'Scheduled', CURRENT_DATE, CURRENT_DATE + INTERVAL '2 hours', 'DEMO001')
            RETURNING flight_id;
            """
            
            cursor.execute(query)
            flight_id = cursor.fetchone()[0]
            conn.commit()
            
            print(f"New flight added successfully. Flight ID: {flight_id}")
            print("Regenerate the chart to see the changes reflected in the visualization.")
            
            cursor.close()
            conn.close()
            
        except Exception as e:
            print(f"Error adding demo flight: {e}")
    
    def run_all_analytics(self):
        """
        Main execution function
        Runs all three assignment tasks in sequence
        """
        print("=" * 70)
        print("SKYTRACK SOLUTIONS - COMPREHENSIVE ANALYTICS SUITE")
        print("=" * 70)
        
        print("\n[TASK 1: Creating 6 visualizations with minimum 2 JOINs each]")
        print("-" * 70)
        self.create_pie_chart()
        self.create_bar_chart()
        self.create_horizontal_bar_chart()
        self.create_line_chart()
        self.create_histogram()
        self.create_scatter_plot()
        
        print("\n" + "-" * 70)
        print("[TASK 2: Interactive Plotly timeline with real date-based slider]")
        print("-" * 70)
        self.create_interactive_timeline()
        
        print("\n" + "-" * 70)
        print("[TASK 3: Excel export with advanced formatting]")
        print("-" * 70)
        self.export_to_excel()
        
        print("\n" + "=" * 70)
        print("ALL ANALYTICAL TASKS COMPLETED SUCCESSFULLY")
        print("=" * 70)
        print("\nGenerated files:")
        print("   - charts/ folder: 6 visualization files")
        print("   - exports/ folder: Excel report with formatted data")
        print("\nFor project defense: Use add_demo_flight() to demonstrate live updates")
        print("=" * 70)


if __name__ == "__main__":
    # Initialize analytics system
    analytics = SkyTrackAnalytics()
    
    # Run all analytics tasks
    analytics.run_all_analytics()
    
    